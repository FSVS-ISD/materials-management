import os
from io import BytesIO
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import logging

from flask import Blueprint, request, jsonify, send_file, g
from flask_jwt_extended import jwt_required
from extensions import db
from models import Material, InRecord, OutRecord
from sqlalchemy import func

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

report_bp = Blueprint('report', __name__)
logger = logging.getLogger(__name__)

# 字型設定，請確保 fonts 目錄及字型檔存在
basedir = os.path.abspath(os.path.dirname(__file__))
FONT_PATH = os.path.join(basedir, '..', 'fonts', 'NotoSansTC-Regular.ttf')

try:
    if os.path.exists(FONT_PATH):
        pdfmetrics.registerFont(TTFont('ChineseFont', FONT_PATH))
        pdfmetrics.registerFontFamily('ChineseFont', normal='ChineseFont', bold='ChineseFont', italic='ChineseFont', boldItalic='ChineseFont')
        logger.info(f"中文字型 'ChineseFont' 載入成功。")
    else:
        logger.error(f"找不到字型檔 {FONT_PATH}。PDF 報表中的中文可能無法正常顯示。")
except Exception as e:
    logger.error(f"註冊中文字型失敗: {e}")

REPORT_TYPE_MAP = {
    'stock_summary': '庫存摘要報表',
    'in_records': '入庫明細查詢',
    'out_records': '出庫明細查詢',
    'low_stock_alert': '低庫存警示報表'
}

def get_report_params(args):
    report_type = args.get('report_type', 'stock_summary')
    query_mode = args.get('query_mode', 'daterange')
    category = args.get('category')
    item_id = args.get('item_id')
    school_dept = args.get('school_dept', '鳳山商工 ****科')

    dt_start, dt_end = None, None
    target_year, target_month = None, None
    
    if report_type == 'low_stock_alert':
        return report_type, category, item_id, school_dept, None, None, None, None

    if query_mode == 'daterange':
        start_date_str = args.get('start_date')
        end_date_str = args.get('end_date')
        if not start_date_str or not end_date_str:
            raise ValueError("在日期範圍模式下，必須提供開始與結束日期。")
        dt_start = datetime.strptime(start_date_str, '%Y-%m-%d').replace(hour=0, minute=0, second=0)
        dt_end = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        target_year, target_month = dt_end.year, dt_end.month
    else:
        year_str = args.get('year')
        month_str = args.get('month')
        if not year_str or not month_str:
            raise ValueError("在月份模式下，必須提供年份與月份。")
        target_year, target_month = int(year_str), int(month_str)
        dt_start = datetime(target_year, target_month, 1)
        dt_end = dt_start + relativedelta(months=1) - timedelta(seconds=1)

    return report_type, category, item_id, school_dept, dt_start, dt_end, target_year, target_month

def calculate_stock_at_date(session, material_id, specific_date):
    try:
        total_in = session.query(func.coalesce(func.sum(InRecord.quantity), 0)).filter(
            InRecord.material_id == material_id,
            InRecord.date < specific_date
        ).scalar()
        total_out = session.query(func.coalesce(func.sum(OutRecord.quantity), 0)).filter(
            OutRecord.material_id == material_id,
            OutRecord.date < specific_date
        ).scalar()
        return total_in - total_out
    except Exception as e:
        logger.error(f"計算物料 {material_id} 在 {specific_date} 的期初庫存時出錯: {e}")
        return 0

def calculate_monthly_io(session, material_id, year, month):
    start_of_month = datetime(year, month, 1)
    end_of_month = start_of_month + relativedelta(months=1)
    monthly_in = session.query(func.coalesce(func.sum(InRecord.quantity), 0)).filter(
        InRecord.material_id == material_id,
        InRecord.date >= start_of_month,
        InRecord.date < end_of_month
    ).scalar()
    monthly_out = session.query(func.coalesce(func.sum(OutRecord.quantity), 0)).filter(
        OutRecord.material_id == material_id,
        OutRecord.date >= start_of_month,
        OutRecord.date < end_of_month
    ).scalar()
    return monthly_in, monthly_out

@report_bp.route('/api/report/preview', methods=['GET'])
@jwt_required()
def report_preview_pdf():
    session = g.db_session()
    try:
        params = get_report_params(request.args)
        report_type, category, item_id, school_dept, dt_start, dt_end, target_year, target_month = params
        query_mode = request.args.get('query_mode', 'daterange')
    except ValueError as e:
        logger.error(f"PDF 報表參數錯誤: {e}")
        return jsonify({'error': str(e)}), 400

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)

    styles = getSampleStyleSheet()
    styleN = ParagraphStyle(name='Normal_Chinese', parent=styles['Normal'], fontName='ChineseFont', fontSize=9, leading=12)
    styleH = ParagraphStyle(name='Heading1_Chinese', parent=styles['Heading1'], fontName='ChineseFont', alignment=TA_CENTER)
    # 新增紅色字體樣式
    styleRed = ParagraphStyle(name='Red_Chinese', parent=styles['Normal'], fontName='ChineseFont', fontSize=9, leading=12, textColor=colors.red)

    elements = []
    report_title = REPORT_TYPE_MAP.get(report_type, report_type)

    # 修正查詢期間顯示問題
    if report_type == 'low_stock_alert':
        # 低庫存警示報表不需要查詢期間
        title_text = f"{school_dept}  {report_title}"
    else:
        if query_mode == 'month':
            query_time_text = f"{target_year}年{target_month}月"
        else:
            query_time_text = f"{dt_start.strftime('%Y/%m/%d')} - {dt_end.strftime('%Y/%m/%d')}"
        title_text = f"{school_dept}  {report_title}  （查詢期間：{query_time_text}）"
    
    elements.append(Paragraph(title_text, styleH))
    elements.append(Spacer(1, 12))

    if report_type == 'stock_summary':
        headers = ["物料編號", "分類", "名稱", "單位", "上月庫存", "本月入庫", "本月出庫", "實際庫存", "安全庫存", "備註/存放點"]
        data = [headers]

        query = session.query(Material)
        if category and category != 'all':
            query = query.filter(Material.category == category)
        if item_id and item_id != 'all':
            query = query.filter(Material.item_id == item_id)
        materials = query.order_by(Material.item_id).all()

        start_of_month = datetime(target_year, target_month, 1)

        for m in materials:
            prev_month_stock = calculate_stock_at_date(session, m.id, start_of_month)
            monthly_in, monthly_out = calculate_monthly_io(session, m.id, target_year, target_month)
            end_of_month_stock = prev_month_stock + monthly_in - monthly_out

            notes_text = m.notes or ''
            is_low_stock = False
            if m.safety_stock > 0 and end_of_month_stock <= m.safety_stock:
                notes_text = "低庫存" if not notes_text else f"低庫存; {notes_text}"
                is_low_stock = True

            # 根據是否為低庫存選擇不同的樣式
            notes_paragraph = Paragraph(notes_text, styleRed if is_low_stock else styleN)

            data.append([
                Paragraph(m.item_id, styleN), Paragraph(m.category, styleN), Paragraph(m.name, styleN),
                Paragraph(m.unit, styleN), prev_month_stock, monthly_in, monthly_out,
                end_of_month_stock, m.safety_stock, notes_paragraph
            ])

        table = Table(data, colWidths=[50, 60, 110, 30, 50, 50, 50, 50, 50, 65], repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('GRID', (0,0), (-1,-1), 0.5, colors.blue),
            ('FONTNAME', (0,0), (-1,-1), 'ChineseFont'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (4,1), (-1,-1), 'RIGHT')
        ]))
        elements.append(table)

        elements.append(Spacer(1, 40))
        current_date_str = datetime.now().strftime("%Y-%m-%d")
        footer_data = [[
            '製表人:', 
            '科主任:', 
            '實習組長:', 
            '實習主任:', 
            f'製表日期: {current_date_str}'
        ]]

        footer_table = Table(footer_data, colWidths=[107] * 5)
        footer_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'ChineseFont'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
        ]))
        elements.append(footer_table)

    elif report_type in ['in_records', 'out_records']:
        is_in_record = report_type == 'in_records'
        model = InRecord if is_in_record else OutRecord
        headers = ["日期", "物料編號", "名稱", "分類", "數量", "來源/部門", "經手人/用途"]
        data = [headers]

        query = session.query(model, Material).join(Material, model.material_id == Material.id)
        if dt_start:
            query = query.filter(model.date >= dt_start)
        if dt_end:
            query = query.filter(model.date <= dt_end)
        if category and category != 'all':
            query = query.filter(Material.category == category)
        if item_id and item_id != 'all':
            query = query.filter(Material.item_id == item_id)
        records = query.order_by(model.date.desc()).all()

        for r, m in records:
            data.append([
                r.date.strftime('%Y-%m-%d'), m.item_id, Paragraph(m.name, styleN), m.category, r.quantity,
                r.source or '' if is_in_record else r.department or '',
                r.handler or '' if is_in_record else r.purpose or '',
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('FONTNAME', (0,0), (-1,-1), 'ChineseFont'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (4,1), (4,-1), 'RIGHT')
        ]))
        elements.append(table)

    elif report_type == 'low_stock_alert':
        headers = ["物料編號", "分類", "名稱", "單位", "安全庫存", "目前庫存", "庫存差距"]
        data = [headers]

        query = session.query(Material).filter(Material.safety_stock > 0)\
                              .filter(Material.current_stock <= Material.safety_stock)

        if category and category != 'all':
            query = query.filter(Material.category == category)
        if item_id and item_id != 'all':
            query = query.filter(Material.item_id == item_id)

        materials = query.order_by(Material.item_id).all()

        for m in materials:
            stock_gap = m.safety_stock - m.current_stock
            data.append([
                m.item_id, m.category, Paragraph(m.name, styleN), m.unit,
                m.safety_stock, m.current_stock, stock_gap
            ])

        table = Table(data, colWidths=[70, 80, 150, 40, 60, 60, 60], repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('FONTNAME', (0,0), (-1,-1), 'ChineseFont'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (4,1), (-1,-1), 'RIGHT'),
            ('TEXTCOLOR', (6, 1), (6, -1), colors.red),
        ]))
        elements.append(table)

    else:
        elements.append(Paragraph("未知的報表類型", styleN))

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=False, download_name=f"{report_type}_preview.pdf", mimetype='application/pdf')

@report_bp.route('/api/report/export_excel', methods=['GET'])
@jwt_required()
def report_export_excel():
    session = g.db_session()
    try:
        params = get_report_params(request.args)
        report_type, category, item_id, school_dept, dt_start, dt_end, target_year, target_month = params
        query_mode = request.args.get('query_mode', 'daterange')
    except ValueError as e:
        logger.error(f"Excel 報表參數錯誤: {e}")
        return jsonify({'error': str(e)}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = REPORT_TYPE_MAP.get(report_type, report_type)

    font_header = Font(bold=True, name='Calibri')
    font_red = Font(color="FF0000", name='Calibri')  # 新增紅色字體
    align_center = Alignment(horizontal='center', vertical='center')
    fill_header = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    # 修正查詢期間顯示問題
    if report_type == 'low_stock_alert':
        # 低庫存警示報表不需要查詢期間
        title_text = f"{school_dept}  {REPORT_TYPE_MAP.get(report_type, report_type)}"
    else:
        if query_mode == 'month':
            query_time_text = f"{target_year}年{target_month}月"
        else:
            query_time_text = f"{dt_start.strftime('%Y/%m/%d')} - {dt_end.strftime('%Y/%m/%d')}"
        title_text = f"{school_dept}  {REPORT_TYPE_MAP.get(report_type, report_type)}  （查詢期間：{query_time_text}）"
    
    ws.append([title_text])
    ws.append([])

    if report_type == 'stock_summary':
        headers = ["物料編號", "分類", "名稱", "單位", "上月庫存", "本月入庫", "本月出庫", "實際庫存", "安全庫存", "備註/存放點"]
        ws.append(headers)

        query = session.query(Material)
        if category and category != 'all':
            query = query.filter(Material.category == category)
        if item_id and item_id != 'all':
            query = query.filter(Material.item_id == item_id)
        materials = query.order_by(Material.item_id).all()

        start_of_month = datetime(target_year, target_month, 1)

        for row_idx, m in enumerate(materials, start=3):  # 從第3行開始（標題在第1行，空行在第2行）
            prev_month_stock = calculate_stock_at_date(session, m.id, start_of_month)
            monthly_in, monthly_out = calculate_monthly_io(session, m.id, target_year, target_month)
            end_of_month_stock = prev_month_stock + monthly_in - monthly_out

            notes_text = m.notes or ''
            is_low_stock = False
            if m.safety_stock > 0 and end_of_month_stock <= m.safety_stock:
                notes_text = "低庫存" if not notes_text else f"低庫存; {notes_text}"
                is_low_stock = True

            ws.append([
                m.item_id, m.category, m.name, m.unit,
                prev_month_stock, monthly_in, monthly_out,
                end_of_month_stock, m.safety_stock, notes_text
            ])
            
            # 如果為低庫存，設置備註/存放點單元格為紅色字體
            if is_low_stock:
                notes_cell = ws.cell(row=row_idx, column=10)  # 第10列是"備註/存放點"
                notes_cell.font = font_red

        footer_start_row = ws.max_row + 3
        ws.cell(row=footer_start_row, column=1, value='製表人:')
        ws.cell(row=footer_start_row, column=3, value='科主任:')
        ws.cell(row=footer_start_row, column=5, value='實習組長:')
        ws.cell(row=footer_start_row, column=7, value='實習主任:')

        current_date_str = datetime.now().strftime("%Y-%m-%d")
        ws.cell(row=footer_start_row + 2, column=1, value=f"製表日期: {current_date_str}")

    elif report_type in ['in_records', 'out_records']:
        is_in_record = report_type == 'in_records'
        model = InRecord if is_in_record else OutRecord
        headers = ["日期", "物料編號", "名稱", "分類", "數量", "來源/部門", "經手人/用途"]
        ws.append(headers)

        query = session.query(model, Material).join(Material, model.material_id == Material.id)
        if dt_start:
            query = query.filter(model.date >= dt_start)
        if dt_end:
            query = query.filter(model.date <= dt_end)
        if category and category != 'all':
            query = query.filter(Material.category == category)
        if item_id and item_id != 'all':
            query = query.filter(Material.item_id == item_id)
        records = query.order_by(model.date.desc()).all()

        for r, m in records:
            ws.append([
                r.date.strftime('%Y-%m-%d'), m.item_id, m.name, m.category, r.quantity,
                r.source or '' if is_in_record else r.department or '',
                r.handler or '' if is_in_record else r.purpose or '',
            ])

    elif report_type == 'low_stock_alert':
        headers = ["物料編號", "分類", "名稱", "單位", "安全庫存", "目前庫存", "庫存差距"]
        ws.append(headers)

        query = session.query(Material).filter(Material.safety_stock > 0)\
                              .filter(Material.current_stock <= Material.safety_stock)

        if category and category != 'all':
            query = query.filter(Material.category == category)
        if item_id and item_id != 'all':
            query = query.filter(Material.item_id == item_id)

        materials = query.order_by(Material.item_id).all()

        for m in materials:
            stock_gap = m.safety_stock - m.current_stock
            ws.append([
                m.item_id, m.category, m.name, m.unit,
                m.safety_stock, m.current_stock, stock_gap
            ])

    else:
        ws.append(["未知的報表類型"])

    # 調整欄寬與樣式
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = fill_header
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for r in ws.iter_rows(min_row=1, max_col=col_idx, min_col=col_idx):
            try:
                if len(str(r[0].value)) > max_length:
                    max_length = len(str(r[0].value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{report_type}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')